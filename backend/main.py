from database import (init_db, add_player as db_add_player, get_players, update_player, get_tier_settings, update_tier_baseline, reset_tier_players, get_tier_for_rating, create_snapshot, undo_last_tournament, get_db_connection, delete_player)
from glicko2 import Player
from fastapi import FastAPI, UploadFile, File, HTTPException
from pydantic import BaseModel
import openpyxl
import io
import difflib
from typing import Optional

app = FastAPI()
players = {}

def reload_players_in_memory():
    global players
    players.clear()
    raw_data = get_players()
    for p in raw_data:
        players[p['name']] = Player(rating=p['rating'], rd=p['rd'], vol=p['vol'])

init_db()
reload_players_in_memory()

current_tournament_games = []

class PlayerInput(BaseModel):
    name: str
    rating: float = 1500.0
    rd: float = 200.0
    vol: float = 0.06

class PlayerUpdate(BaseModel):
    new_name: Optional[str] = None
    rating: Optional[float] = None
    rd: Optional[float] = None
    vol: Optional[float] = None
    tier: Optional[str] = None

class GameInput(BaseModel):
    player1_name: str
    player2_name: str
    result: float

class TierUpdateInput(BaseModel):
    group_tier: str
    rating: float
    rd: float
    vol: float

class TierResetInput(BaseModel):
    group_tier: str

def resolve_player(excel_name: str, db_players: list):
    excel_name_lower = excel_name.lower().strip()
    if excel_name in db_players:
        return excel_name, []

    word_matches = [
        db_p for db_p in db_players 
        if excel_name_lower in [part.lower() for part in db_p.split()]
    ]
    if len(word_matches) == 1:
        return word_matches[0], []
    elif len(word_matches) > 1:
        return None, word_matches

    suggestions = set()
    cutoff = 0.6 if len(excel_name) < 5 else 0.45 
    for db_p in db_players:
        db_first_name = db_p.split()[0].lower()
        ratio_first = difflib.SequenceMatcher(None, excel_name_lower, db_first_name).ratio()
        if ratio_first >= cutoff:
            suggestions.add(db_p)

    sorted_suggestions = sorted(
        list(suggestions),
        key=lambda x: difflib.SequenceMatcher(None, excel_name_lower, x.split()[0].lower()).ratio(),
        reverse=True
    )[:3]

    if sorted_suggestions:
        return None, sorted_suggestions
    return None, []

def resolve_winner(excel_winner: str, p1: str, p2: str):
    if excel_winner == "Draw" or excel_winner in [p1, p2]:
        return excel_winner
    closest = difflib.get_close_matches(excel_winner, [p1, p2], n=1, cutoff=0.4)
    return closest[0] if closest else excel_winner

@app.get("/players")
def get_all_players():
    return get_players()

@app.post("/players")
def create_player(player: PlayerInput):
    db_add_player(player.name, player.rating, player.rd, player.vol)
    players[player.name] = Player(rating=player.rating, rd=player.rd, vol=player.vol)
    return {"message": f"Player {player.name} added successfully"}

@app.delete("/players/{player_name}")
def api_delete_player(player_name: str):
    success = delete_player(player_name)
    if not success:
        raise HTTPException(status_code=404, detail="Player not found")
        
    if player_name in players:
        del players[player_name] # Remove from RAM instantly
        
    return {"message": f"Player {player_name} deleted successfully"}

@app.put("/players/{player_name}")
def api_edit_player(player_name: str, update_data: PlayerUpdate):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM players WHERE name = ?", (player_name,))
    if not cursor.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Player not found")

    updates = []
    params = []
    
    if update_data.new_name is not None:
        updates.append("name = ?")
        params.append(update_data.new_name)
    if update_data.rating is not None:
        updates.append("rating = ?")
        params.append(update_data.rating)
    if update_data.rd is not None:
        updates.append("rd = ?")
        params.append(update_data.rd)
    if update_data.vol is not None:
        updates.append("vol = ?")
        params.append(update_data.vol)
    if update_data.tier is not None:
        updates.append("group_tier = ?")
        params.append(update_data.tier)

    if not updates:
        conn.close()
        return {"message": "No data provided to update"}

    params.append(player_name)
    query = f"UPDATE players SET {', '.join(updates)} WHERE name = ?"
    
    cursor.execute(query, tuple(params))
    conn.commit()
    conn.close()
    
    # Refresh RAM so the math uses the edited stats!
    reload_players_in_memory() 
    
    return {"message": f"Player {player_name} updated successfully"}

@app.post("/games")
def add_game(game: GameInput):
    if game.player1_name not in players or game.player2_name not in players:
        raise HTTPException(status_code=404, detail="One or both players not found")
    
    if game.player1_name == game.player2_name:
        raise HTTPException(status_code=400, detail="Cannot play against yourself")
        
    current_tournament_games.append(game)
    return {"status": "Game added", "queue_length": len(current_tournament_games)}

@app.post("/tournament/process")
def process_tournament():
    global current_tournament_games
    if not current_tournament_games: return {"message": "No games to process"}
    create_snapshot()
    
    # 1. FREEZE THE STATS: Take a snapshot of everyone's rating before doing any math!
    pre_tournament_stats = {
        name: {"rating": p.rating, "rd": p.rd, "vol": p.vol} 
        for name, p in players.items()
    }
    
    changes = {}
    for game in current_tournament_games:
        result = game.result
        p1 = game.player1_name
        p2 = game.player2_name

        if p1 not in changes: changes[p1] = []
        changes[p1].append((p2, result))

        if p2 not in changes: changes[p2] = []
        changes[p2].append((p1, 1.0 - result))

    response_data = []
    for name, matches in changes.items():
        if name not in players: continue
        
        player_obj = players[name]
        old_rating = player_obj.rating
        old_rd = player_obj.rd
        old_vol = player_obj.vol
        
        opponent_ratings = []
        opponent_rds = []
        outcomes = []

        for opponent_name, outcome in matches:
            if opponent_name not in pre_tournament_stats: continue
            
            # 2. READ THE FROZEN STATS: Use the pre-tournament snapshot, not the live object!
            frozen_opp = pre_tournament_stats[opponent_name]
            opponent_ratings.append(float(frozen_opp["rating"]))
            opponent_rds.append(float(frozen_opp["rd"]))
            outcomes.append(float(outcome))
        
        player_obj.update_player(opponent_ratings, opponent_rds, outcomes)

        new_tier = get_tier_for_rating(player_obj.rating)

        update_player(
            name=name, 
            new_rating=player_obj.rating, 
            new_rd=player_obj.rd, 
            new_vol=player_obj.vol, 
            new_tier=new_tier
        )

        response_data.append({
            "name": name,
            "old_rating": round(old_rating, 2),
            "new_rating": round(player_obj.rating, 2),
            "old_rd": round(old_rd, 2),
            "new_rd": round(player_obj.rd, 2),
            "old_vol": round(old_vol, 4),
            "new_vol": round(player_obj.vol, 4),
            "new_tier": new_tier
        })
        
    current_tournament_games = []
    
    return response_data

@app.get("/tiers")
def get_tiers():
    return get_tier_settings()

@app.post("/tiers/update")
def update_tier(tier: TierUpdateInput):
    update_tier_baseline(tier.group_tier, tier.rating, tier.rd, tier.vol)
    return {"message": f"{tier.group_tier} baseline updated successfully."}

@app.post("/tiers/reset")
def reset_tier(reset: TierResetInput):
    reset_tier_players(reset.group_tier)
    reload_players_in_memory() # Ensure RAM matches the reset DB
    return {"message": f"All {reset.group_tier} players reset to baseline."}

@app.post("/tournament/undo")
def undo_tournament():
    success = undo_last_tournament()
    if success:
        reload_players_in_memory() # Sync the RAM back to the old stats!
        return {"message": "Tournament successfully undone. Ratings restored."}
    else:
        raise HTTPException(status_code=400, detail="No tournament history found to undo.")
    
@app.post("/upload_tournament_excel/")
async def upload_tournament_excel(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents))
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM players")
    
    try:
        valid_players = [row['name'] for row in cursor.fetchall()]
    except TypeError:
        valid_players = [row[0] for row in cursor.fetchall()]
        
    conn.close()

    valid_games = []
    conflicts = []
    errors = []

    # 3. Parse the rows (Starting at Row 4)
    for row_idx in range(4, sheet.max_row + 1):
        winner_val = sheet.cell(row=row_idx, column=2).value
        match_val = sheet.cell(row=row_idx, column=3).value

        if not match_val:
            continue 

        winner_raw = str(winner_val).strip() if winner_val else ""
        match_str = str(match_val).strip()

        if " vs " not in match_str:
            errors.append(f"Row {row_idx}: Bad match format -> '{match_str}'")
            continue

        parts = match_str.split(" vs ")
        p1_raw, p2_raw = parts[0].strip(), parts[1].strip()

        corrected_winner_raw = resolve_winner(winner_raw, p1_raw, p2_raw)

        p1_resolved, p1_suggestions = resolve_player(p1_raw, valid_players)
        p2_resolved, p2_suggestions = resolve_player(p2_raw, valid_players)

        # 4. Route to Valid Games or Conflicts
        if p1_resolved and p2_resolved:
            final_winner = p1_resolved if corrected_winner_raw == p1_raw else (p2_resolved if corrected_winner_raw == p2_raw else "Draw")
            valid_games.append({
                "row": row_idx,
                "player1": p1_resolved,
                "player2": p2_resolved,
                "winner": final_winner,
                "score": 1.0 if final_winner == p1_resolved else (0.5 if final_winner == "Draw" else 0.0)
            })
        else:
            conflicts.append({
                "row": row_idx,
                "raw_match": match_str,
                "p1_target": p1_raw,
                "p1_suggestions": p1_suggestions,
                "p2_target": p2_raw,
                "p2_suggestions": p2_suggestions,
                "raw_winner": corrected_winner_raw
            })

    return {
        "message": "File processed",
        "valid_games": valid_games,
        "conflicts": conflicts,
        "errors": errors
    }